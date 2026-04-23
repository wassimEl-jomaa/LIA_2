# app/services/impexp.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Literal, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import io
import csv
import json

# YAML kräver dependency: pyyaml
try:
    import yaml  # type: ignore
except Exception:
    yaml = None

from openpyxl import Workbook, load_workbook

from ..models import Requirement, TestCase  # anpassa till dina modeller

ExportFormat = Literal["csv", "xlsx", "json", "yaml"]
DeclaredImportFormat = Literal["auto", "csv", "xlsx", "json", "yaml"]
ImportMode = Literal["create_only", "upsert_external_id"]
OnErrorMode = Literal["stop", "continue"]


CANON_REQ_FIELDS = ["external_id", "title", "description", "priority", "status", "tags", "source"]


# --- 1) Column mapping (auto + manuellt) ---

SYNONYMS: dict[str, list[str]] = {
    "external_id": ["external_id", "id", "req_id", "requirement_id", "key", "identifier"],
    "title": ["title", "name", "summary", "requirement", "requirement_title"],
    "description": ["description", "desc", "details", "text", "body"],
    "priority": ["priority", "prio", "importance"],
    "status": ["status", "state"],
    "tags": ["tags", "labels", "label", "tag", "categories", "category"],
    "source": ["source", "origin", "system"],
}


def _norm_header(s: str) -> str:
    return "".join(ch.lower() for ch in (s or "").strip() if ch.isalnum() or ch in ["_", "-", " "]).replace("-", "_").replace(" ", "_")


def guess_mapping(headers: list[str]) -> dict[str, str]:
    """
    Returnerar mapping {canon_field: header_name_in_file}
    """
    norm_to_raw = {_norm_header(h): h for h in headers}
    out: dict[str, str] = {}

    for canon, syns in SYNONYMS.items():
        for s in syns:
            key = _norm_header(s)
            if key in norm_to_raw:
                out[canon] = norm_to_raw[key]
                break
    return out


def apply_mapping(row: dict[str, Any], mapping: dict[str, str]) -> dict[str, Any]:
    """
    mapping: {canon_field: header_in_row}
    """
    out: dict[str, Any] = {}
    for canon in CANON_REQ_FIELDS:
        src = mapping.get(canon)
        out[canon] = row.get(src) if src else None
    return out


# --- 2) Parse files to rows ---

def detect_format(filename: str, declared: DeclaredImportFormat) -> str:
    if declared != "auto":
        return declared
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        return "csv"
    if fn.endswith(".xlsx") or fn.endswith(".xlsm") or fn.endswith(".xltx"):
        return "xlsx"
    if fn.endswith(".json"):
        return "json"
    if fn.endswith(".yaml") or fn.endswith(".yml"):
        return "yaml"
    # fallback
    return "csv"


def parse_csv_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    rows = [dict(r) for r in reader]
    headers = list(reader.fieldnames or [])
    return rows, headers


def parse_xlsx_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers_row = next(rows_iter, None)
    if not headers_row:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    out: list[dict[str, Any]] = []
    for r in rows_iter:
        d: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        # ignorera helt tomma rader
        if any(v not in [None, "", " "] for v in d.values()):
            out.append(d)
    return out, headers


def parse_json_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    data = json.loads(content.decode("utf-8"))
    # stöd: { requirements: [...] } eller direkt [...]
    if isinstance(data, dict) and isinstance(data.get("requirements"), list):
        rows = data["requirements"]
    elif isinstance(data, list):
        rows = data
    else:
        raise ValueError("JSON must be an array or an object with 'requirements' array")
    if not rows:
        return [], []
    headers = sorted({k for r in rows if isinstance(r, dict) for k in r.keys()})
    return [dict(r) for r in rows], headers


def parse_yaml_bytes(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    if yaml is None:
        raise ValueError("pyyaml is not installed. Add dependency: pyyaml")
    data = yaml.safe_load(content.decode("utf-8"))
    if isinstance(data, dict) and isinstance(data.get("requirements"), list):
        rows = data["requirements"]
    elif isinstance(data, list):
        rows = data
    else:
        raise ValueError("YAML must be a list or have 'requirements' list")
    if not rows:
        return [], []
    headers = sorted({k for r in rows if isinstance(r, dict) for k in r.keys()})
    return [dict(r) for r in rows], headers


def parse_upload(content: bytes, filename: str, declared: DeclaredImportFormat) -> tuple[list[dict[str, Any]], list[str], str]:
    fmt = detect_format(filename, declared)
    if fmt == "csv":
        rows, headers = parse_csv_bytes(content)
    elif fmt == "xlsx":
        rows, headers = parse_xlsx_bytes(content)
    elif fmt == "json":
        rows, headers = parse_json_bytes(content)
    elif fmt == "yaml":
        rows, headers = parse_yaml_bytes(content)
    else:
        raise ValueError(f"Unknown format: {fmt}")
    return rows, headers, fmt


# --- 3) Normalize + validate ---

def _to_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _normalize_tags(v: Any) -> list[str]:
    if v is None:
        return []
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    s = str(v).strip()
    if not s:
        return []
    # "a,b,c"
    parts = [p.strip() for p in s.split(",")]
    return [p for p in parts if p]


def normalize_requirement(raw: dict[str, Any]) -> dict[str, Any]:
    ext = _to_str(raw.get("external_id"))
    title = _to_str(raw.get("title"))
    if not title:
        raise ValueError("Missing title")

    prio = _to_str(raw.get("priority"))
    if prio:
        prio = prio.lower()

    status = _to_str(raw.get("status"))
    if status:
        status = status.lower()

    return {
        "external_id": ext,
        "title": title,
        "description": _to_str(raw.get("description")),
        "priority": prio,
        "status": status,
        "tags": _normalize_tags(raw.get("tags")),
        "source": _to_str(raw.get("source")),
    }


# --- 4) Import orchestrator ---

async def import_requirements_from_upload(
    db: AsyncSession,
    project_id: int,
    upload,  # UploadFile
    declared_format: DeclaredImportFormat,
    mode: ImportMode,
    dry_run: bool,
    on_error: OnErrorMode,
    mapping: Optional[dict[str, str]],
    user_id: int,
) -> dict[str, Any]:
    content = await upload.read()
    rows, headers, fmt = parse_upload(content, upload.filename or "", declared_format)

    auto_mapping = guess_mapping(headers)
    effective_mapping = mapping or auto_mapping

    created = 0
    updated = 0
    skipped = 0
    errors: list[dict[str, Any]] = []
    preview: list[dict[str, Any]] = []

    # prefetch existing by external_id if needed
    existing_by_ext: dict[str, Requirement] = {}
    if mode == "upsert_external_id":
        ext_ids = []
        for r in rows:
            mapped = apply_mapping(r, effective_mapping)
            ext = _to_str(mapped.get("external_id"))
            if ext:
                ext_ids.append(ext)
        if ext_ids:
            q = await db.execute(
                select(Requirement).where(
                    Requirement.project_id == project_id,
                    Requirement.external_id.in_(ext_ids),
                )
            )
            for req in q.scalars().all():
                if req.external_id:
                    existing_by_ext[req.external_id] = req

    for idx, r in enumerate(rows, start=2 if fmt in ["csv", "xlsx"] else 1):
        # idx: radnummer för felrapport (csv/xlsx räknar header som rad 1)
        try:
            mapped = apply_mapping(r, effective_mapping)
            norm = normalize_requirement(mapped)
            preview.append(norm)

            ext = norm.get("external_id")
            if mode == "create_only":
                if ext and ext in existing_by_ext:
                    skipped += 1
                    continue
                if dry_run:
                    created += 1
                    continue
                db.add(
                    Requirement(
                        project_id=project_id,
                        external_id=ext,
                        title=norm["title"],
                        description=norm.get("description"),
                        priority=norm.get("priority"),
                        status=norm.get("status"),
                        tags=norm.get("tags"),
                        source=norm.get("source"),
                        created_by_user_id=user_id,  # om du har fältet
                    )
                )
                created += 1

            else:  # upsert_external_id
                if ext and ext in existing_by_ext:
                    if dry_run:
                        updated += 1
                        continue
                    obj = existing_by_ext[ext]
                    obj.title = norm["title"]
                    obj.description = norm.get("description")
                    obj.priority = norm.get("priority")
                    obj.status = norm.get("status")
                    obj.tags = norm.get("tags")
                    obj.source = norm.get("source")
                    updated += 1
                else:
                    if dry_run:
                        created += 1
                        continue
                    db.add(
                        Requirement(
                            project_id=project_id,
                            external_id=ext,
                            title=norm["title"],
                            description=norm.get("description"),
                            priority=norm.get("priority"),
                            status=norm.get("status"),
                            tags=norm.get("tags"),
                            source=norm.get("source"),
                            created_by_user_id=user_id,  # om du har fältet
                        )
                    )
                    created += 1

        except Exception as e:
            errors.append({"row": idx, "message": str(e)})
            if on_error == "stop":
                break

    if not dry_run and (created or updated):
        await db.commit()

    return {
        "format": fmt,
        "dry_run": dry_run,
        "headers": headers,
        "auto_mapping": auto_mapping,
        "effective_mapping": effective_mapping,
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "preview": preview[:20],  # preview max 20 rader
    }


# --- 5) Exporters ---

async def export_requirements_bytes(db: AsyncSession, project_id: int, fmt: ExportFormat) -> tuple[bytes, str]:
    q = await db.execute(select(Requirement).where(Requirement.project_id == project_id).order_by(Requirement.id.asc()))
    reqs = q.scalars().all()

    rows = []
    for r in reqs:
        rows.append({
            "external_id": getattr(r, "external_id", None),
            "title": r.title,
            "description": getattr(r, "description", None),
            "priority": getattr(r, "priority", None),
            "status": getattr(r, "status", None),
            "tags": getattr(r, "tags", None),
            "source": getattr(r, "source", None),
        })

    return _export_rows(rows, fmt, sheet_name="requirements")


async def export_test_cases_bytes(db: AsyncSession, project_id: int, fmt: ExportFormat) -> tuple[bytes, str]:
    q = await db.execute(select(TestCase).where(TestCase.project_id == project_id).order_by(TestCase.id.asc()))
    cases = q.scalars().all()

    rows = []
    for tc in cases:
        rows.append({
            "id": tc.id,
            "title": getattr(tc, "title", None),
            "description": getattr(tc, "description", None),
            "preconditions": getattr(tc, "preconditions", None),
            "steps": getattr(tc, "steps", None),
            "expected_result": getattr(tc, "expected_result", None),
            "requirement_id": getattr(tc, "requirement_id", None),
        })

    return _export_rows(rows, fmt, sheet_name="test_cases")


def _export_rows(rows: list[dict[str, Any]], fmt: ExportFormat, sheet_name: str) -> tuple[bytes, str]:
    if fmt == "json":
        return json.dumps(rows, ensure_ascii=False, indent=2).encode("utf-8"), "application/json"

    if fmt == "yaml":
        if yaml is None:
            raise ValueError("pyyaml is not installed. Add dependency: pyyaml")
        return yaml.safe_dump(rows, sort_keys=False, allow_unicode=True).encode("utf-8"), "text/yaml"

    if fmt == "csv":
        out = io.StringIO()
        headers = sorted({k for r in rows for k in r.keys()}) if rows else []
        w = csv.DictWriter(out, fieldnames=headers)
        w.writeheader()
        for r in rows:
            # CSV: listas/dict -> json str
            rr = {}
            for k, v in r.items():
                if isinstance(v, (list, dict)):
                    rr[k] = json.dumps(v, ensure_ascii=False)
                else:
                    rr[k] = v
            w.writerow(rr)
        return out.getvalue().encode("utf-8"), "text/csv"

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        headers = sorted({k for r in rows for k in r.keys()}) if rows else []
        ws.append(headers)
        for r in rows:
            row_vals = []
            for h in headers:
                v = r.get(h)
                if isinstance(v, (list, dict)):
                    row_vals.append(json.dumps(v, ensure_ascii=False))
                else:
                    row_vals.append(v)
            ws.append(row_vals)

        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    raise ValueError(f"Unsupported format: {fmt}")
